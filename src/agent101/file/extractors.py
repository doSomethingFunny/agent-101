"""文件内容抽取器。

提供针对不同类型的文件的内容抽取与摘要生成能力。

Functions:
    extract_pdf_summary: 提取PDF文本并生成摘要。
    extract_word_text: 提取Word(docx)文档纯文本。
    summarize_text: 对任意长文本生成简要摘要。
    extract_excel_overview: 提取Excel工作簿的表结构与示例行。
"""

from __future__ import annotations

from typing import Any, Dict, List, Optional

from docx import Document  # python-docx
from openpyxl import load_workbook

from ..logger import get_logger
from ..llm import LLMClient
from ..research.pdf_parser import extract_text_from_pdf, chunk_text, summarize_chunks


logger = get_logger(__name__)


def extract_pdf_summary(path: str, chunk_chars: int = 3000) -> Dict[str, Any]:
    """提取PDF文本与摘要。

    Args:
        path: PDF 文件路径。
        chunk_chars: 分块摘要的每块字符数。

    Returns:
        Dict[str, Any]: 包含 `text_excerpt` 与 `summary` 字段。

    Raises:
        Exception: 当解析失败时抛出异常。
    """

    text = extract_text_from_pdf(path)
    chunks = chunk_text(text, max_chars=chunk_chars)
    summary = summarize_chunks(chunks)
    return {
        "text_excerpt": text[:1000],
        "summary": summary,
    }


def extract_word_text(path: str) -> str:
    """提取 Word(docx) 文档纯文本。

    Args:
        path: DOCX 文件路径。

    Returns:
        str: 合并的段落文本。
    """

    doc = Document(path)
    parts: List[str] = []
    for para in doc.paragraphs:
        t = (para.text or "").strip()
        if t:
            parts.append(t)
    return "\n".join(parts).strip()


def summarize_text(text: str, prompt: Optional[str] = None) -> str:
    """对任意文本生成简要中文摘要。

    Args:
        text: 输入长文本。
        prompt: 可选系统指令。

    Returns:
        str: 摘要文本（中文）。
    """

    llm = LLMClient()
    system = prompt or "你是严谨的文档总结助手，请用中文概述要点，保留结构化条目。"
    messages = [
        {"role": "system", "content": system},
        {"role": "user", "content": text[:8000]},
    ]
    return llm.chat(messages)


def extract_excel_overview(path: str, max_rows: int = 5) -> Dict[str, Any]:
    """提取 Excel 工作簿结构与示例行。

    Args:
        path: Excel 文件路径。
        max_rows: 每个工作表的示例行上限。

    Returns:
        Dict[str, Any]: 包含 `sheets` 列表，每个元素含 `name`、`headers`、`rows`。
    """

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    sheets: List[Dict[str, Any]] = []

    for ws in wb.worksheets:
        # 提取表头（第一行）与示例行
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = list(next(rows_iter) or [])
        except StopIteration:
            headers = []
        samples: List[List[Any]] = []
        for i, row in enumerate(rows_iter):
            if i >= max_rows:
                break
            samples.append(list(row or []))
        sheets.append({"name": ws.title, "headers": headers, "rows": samples})

    return {"sheets": sheets}